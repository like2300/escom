# Create your views here.
from django.shortcuts import render,redirect,get_object_or_404
from database.models import *
from .models import *
from django.http import HttpResponse,HttpResponseRedirect
# from datetime import datetime,date,time
# from emploieDutemps.models import EmploieDuTemps,Jour
# from django.db.models import Q
# from heures_cours.models import PresenceCours
from django.db.models import Count, Sum, Avg,F, FloatField, ExpressionWrapper
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login as auth_login,logout
from django.http import HttpResponseForbidden

# from .utils import *

# from django.shortcuts import render,redirect,get_object_or_404
# # # from dataBase.models import *
from django.contrib import messages
# from django.http import HttpResponse
# from datetime import date,time,datetime
# # from django.utils import timezone
# # # from django.contrib.auth import authenticate,login,logout
from transactions.utils import *
# # from django.conf import settings

from django.db.models import Q
from .models import Transaction, Notes, Classe, Matiere_Programmer
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from .models import Notes, Etudiant, Matiere_Programmer, Classe, AnneeAcademique, Semestre

from django.template.loader import render_to_string
from django.http import HttpResponse
import os
from datetime import datetime
from io import BytesIO
from xhtml2pdf import pisa

def generer_bulletin_pdf(etudiant, semestre, annee_academique):
    """
    Génère un bulletin de notes en PDF
    """
    # Calculer les moyennes
    resultat = calcul_moyenne_simple(etudiant, semestre, annee_academique)
    print(resultat)
    # Déterminer la mention
    moyenne_generale = resultat['moyenne_generale']
    if moyenne_generale >= 16:
        mention = "TRÈS BIEN"
    elif moyenne_generale >= 14:
        mention = "BIEN"
    elif moyenne_generale >= 12:
        mention = "ASSEZ BIEN"
    elif moyenne_generale >= 10:
        mention = "PASSABLE"
    else:
        mention = "INSUFFISANT"
    
    # Préparer les données pour le template
    context = {
        'transaction':Transaction.objects.get(etudiant=etudiant),
        'etudiant': etudiant,
        'semestre': semestre,
        'annee_academique': annee_academique,
        'matieres': resultat['matieres'],
        'moyenne_generale': moyenne_generale,
        'total_coefficients': resultat['total_coefficients'],
        'mention': mention,
        'somme_notes_ponderees': sum(details['ponderee'] for details in resultat['matieres'].values())
    }
    
    # Générer le HTML
    html_content = render_to_string('pages/pdf/bulletin.html', context)
    
    # Options pour pdfkit
    options = {
        'page-size': 'A4',
        'margin-top': '0.5in',
        'margin-right': '0.5in',
        'margin-bottom': '0.5in',
        'margin-left': '0.5in',
        'encoding': "UTF-8",
        'no-outline': None,
        'enable-local-file-access': None
    }
    
    try:
        # Générer le HTML
        html_content = render_to_string('pages/pdf/bulletin.html', context)
        
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html_content.encode("UTF-8")), result)
        
        if not pdf.err:
            # Créer la réponse HTTP
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            filename = f"bulletin_{etudiant.utilisateur.matricule}_{semestre.nom_semestre}_{datetime.now().strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'inline; filename="{filename}"'
            return response
        else:
            return HttpResponse(f"Erreur lors de la génération du PDF: {pdf.err}")
            
    except Exception as e:
        # Gérer l'erreur (retourner une réponse d'erreur ou logger)
        return HttpResponse(f"Erreur lors de la génération du PDF: {str(e)}")

def calcul_moyenne_simple(etudiant, semestre, annee_academique):
    """
    Version adaptée pour le template PDF
    """
    from .models import EmploieDuTemps, Matiere_Programmer, Notes
    
    resultats = {}
    total_pondere = 0
    total_coefficients = 0
    transaction=Transaction.objects.get(etudiant=etudiant)
    # Récupérer les matières de l'emploi du temps
    # matieres_emploi = EmploieDuTemps.objects.filter(
    #     classe=transaction.classe,
        # semestre=semestre,
        # annee_academique=annee_academique
    # ).values_list('matiere_id', flat=True).distinct()
    matieres_emploi=Matiere_Programmer.objects.filter(classe=transaction.classe,
         semestre=semestre,
        annee_academique=annee_academique
        )
    
    for matiere_id in matieres_emploi:
        try:
            matiere = Matiere_Programmer.objects.get(
                id=matiere_id.id,
                semestre=semestre,
                annee_academique=annee_academique,
                classe=transaction.classe
            )
            
            notes = Notes.objects.filter(
                etudiant=etudiant,
                matiere=matiere,
                semestre=semestre,
                annee_academique=annee_academique
            )
            
            note_devoir = notes.filter(evaluation='devoir').first()
            note_session = notes.filter(evaluation='session').first()
            
            note_devoir_val = note_devoir.note if note_devoir else 0
            note_session_val = note_session.note if note_session else 0
            
            # Calcul de la moyenne (30% devoir + 70% session)
            moyenne_matiere = (0.3 * note_devoir_val) + (0.7 * note_session_val)
            moyenne_ponderee = moyenne_matiere * matiere.coefficient
            
            resultats[matiere.nom] = {
                'coefficient': matiere.coefficient,
                'devoir': note_devoir_val,
                'session': note_session_val,
                'moyenne': round(moyenne_matiere, 2),
                'ponderee': round(moyenne_ponderee, 2)
            }
            
            total_pondere += moyenne_ponderee
            total_coefficients += matiere.coefficient
            
        except Matiere_Programmer.DoesNotExist:
            continue
    
    moyenne_generale = total_pondere / total_coefficients if total_coefficients > 0 else 0
    
    return {
        'matieres': resultats,
        'moyenne_generale': round(moyenne_generale, 2),
        'total_coefficients': total_coefficients
    }



def telecharger_bulletin_pdf(request, etudiant_id, semestre_id):
    etudiant = get_object_or_404(Etudiant, id=etudiant_id)
    semestre = get_object_or_404(Semestre, id=semestre_id)
    annee_academique = get_object_or_404(AnneeAcademique, id=1)
    
    return generer_bulletin_pdf(etudiant, semestre, annee_academique)


































@admin_decorators
def menu_saisie(request, classe_id):
    """
    Fonction unifiée pour afficher ET enregistrer les notes
    - Vérifie que les étudiants sont inscrits via la table Transaction
    - Affiche seulement la sélection de matière
    """
    
    # Récupérer la classe

    classe = Classe.objects.get(id=classe_id)
    # Récupérer les données de base
    matieres = Matiere_Programmer.objects.filter(classe=classe)
    # print(matieres.semestre.nom_semestre)
    semestre_actuel = Semestre.objects.filter(
    matiere_programmer__classe=classe
    ).distinct().first()

    

    # Déterminer automatiquement l'année et le semestre
    annee_academique = AnneeAcademique.objects.get(id=1)
    
    # Récupérer seulement les étudiants inscrits (qui ont une transaction pour cette année)
    etudiants_inscrits = Etudiant.objects.filter(
        # transaction__annee_academique=annee_academique,
        transaction__classe=classe
    ).order_by('username').distinct()
    print(etudiants_inscrits)
    
    # Si POST, traiter l'enregistrement des notes
    if request.method == 'POST':
        try:
            # Récupérer la matière sélectionnée
            matiere_id = request.POST.get('matiere')
            evaluation_type = request.POST.get('evaluation')
            
            if not matiere_id or not evaluation_type:
                messages.error(request, 'Veuillez sélectionner une matière et un type d\'évaluation.')
            else:
                matiere_programmee = Matiere_Programmer.objects.get(id=matiere_id)
                
                with transaction.atomic():
                    notes_enregistrees = 0
                    erreurs = []
                    
                    # Parcourir tous les étudiants inscrits
                    for etudiant in etudiants_inscrits:
                        note_field_name = f'note_{etudiant.id}'
                        note_value = request.POST.get(note_field_name)
                    # Si une note a été saisie pour cet étudiant
                        if note_value and note_value.strip():
                            try:
                                note_value = float(note_value)
                                
                                # Vérifier si la note est dans la plage autorisée
                                if note_value < 0 or note_value > 20:
                                    erreurs.append(f"Note invalide pour {etudiant.nom} {etudiant.prenom}: {note_value}")
                                    continue
                                
                                # Vérifier si une note existe déjà pour cette combinaison
                                note_existante = Notes.objects.filter(
                                    annee_academique=annee_academique,
                                    etudiant=etudiant,
                                    semestre=semestre_actuel,
                                    evaluation=evaluation_type,
                                    matiere=matiere_programmee,
                                    classe=classe
                                ).first()
                                
                                if note_existante:
                                    # Mettre à jour la note existante
                                    note_existante.note = note_value
                                    note_existante.save()
                                else:
                                    # Créer une nouvelle note
                                    Notes.objects.create(
                                        annee_academique=annee_academique,
                                        etudiant=etudiant,
                                        semestre=semestre_actuel,
                                        evaluation=evaluation_type,
                                        matiere=matiere_programmee,
                                        classe=classe,
                                        note=note_value
                                    )
                                
                                notes_enregistrees += 1
                                
                            except ValueError:
                                erreurs.append(f"Note invalide pour {etudiant.nom} {etudiant.prenom}: {note_value}")
                            except Exception as e:
                                erreurs.append(f"Erreur pour {etudiant.nom} {etudiant.prenom}: {str(e)}")
                    
                    # Préparer le message de résultat
                    if notes_enregistrees > 0:
                        message = f"{notes_enregistrees} note(s) enregistrée(s) avec succès."
                        if erreurs:
                            message += f" Erreurs: {', '.join(erreurs[:3])}"
                        messages.success(request, message)
                    else:
                        if erreurs:
                            messages.error(request, "Aucune note enregistrée. Erreurs: " + ", ".join(erreurs[:3]))
                        else:
                            messages.warning(request, "Aucune note n'a été saisie.")
                
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'enregistrement: {str(e)}')
    
    # Préparer les données pour l'affichage
    etudiants_avec_notes = []
    for etudiant in etudiants_inscrits:
        # Vérifier si l'étudiant a des notes existantes pour pré-remplir
        note_existante = None
        if request.method == 'POST' and request.POST.get('matiere') and request.POST.get('evaluation'):
            matiere_id = request.POST.get('matiere')
            evaluation_type = request.POST.get('evaluation')
            matiere_programmee = Matiere_Programmer.objects.get(id=matiere_id)
            
            note_existante = Notes.objects.filter(
                annee_academique=annee_academique,
                etudiant=etudiant,
                semestre=semestre_actuel,
                evaluation=evaluation_type,
                matiere=matiere_programmee,
                classe=classe
            ).first()
        
        etudiants_avec_notes.append({
            'etudiant': etudiant,
            'note_existante': note_existante.note if note_existante else None
        })
    
    context = {
        'classe': classe,
        'matieres': matieres,
        'etudiants_avec_notes': etudiants_avec_notes,
        'annee_academique': annee_academique,
        'semestre_actuel': semestre_actuel,
        'total_etudiants': etudiants_inscrits.count(),
        'etudiants_inscrits': etudiants_inscrits,
    }
    
    return render(request, 'pages/notes/saisie_notes.html', context)

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import Workbook

def telecharger_notes_classe_excel(request, classe_id):
    """
    Télécharge les notes de la classe au format Excel avec un style épuré
    """
    classe = get_object_or_404(Classe, id=classe_id)
    transactions = Transaction.objects.filter(classe=classe).select_related('etudiant').order_by('etudiant__utilisateur__nom')
    matieres = Matiere_Programmer.objects.filter(classe=classe).distinct()

    # Créer le classeur Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f'Notes - {classe.nom}'

    # Définir la police pour l'entête
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # Définir les en-têtes
    headers = ["Matricule", "Nom", "Prénom"]
    for matiere in matieres:
        headers.extend([f"{matiere.nom} (Devoir)", f"{matiere.nom} (Session)", f"{matiere.nom} (Moy)"])
    
    # Écrire les en-têtes
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Écrire les données
    for row_num, transaction in enumerate(transactions, 2):
        etudiant = transaction.etudiant
        
        # Remplir les infos de base
        sheet.cell(row=row_num, column=1, value=etudiant.utilisateur.matricule).border = border
        sheet.cell(row=row_num, column=2, value=etudiant.utilisateur.nom).border = border
        sheet.cell(row=row_num, column=3, value=etudiant.utilisateur.prenom).border = border
        
        col_offset = 4
        
        # Remplir les notes par matière
        for matiere in matieres:
            # Récupérer les notes
            note_devoir = Notes.objects.filter(
                etudiant=etudiant, 
                matiere=matiere, 
                evaluation='devoir'
            ).first()
            note_session = Notes.objects.filter(
                etudiant=etudiant, 
                matiere=matiere, 
                evaluation='session'
            ).first()
            
            # Calculer la moyenne
            moyenne = None
            if note_devoir and note_session:
                moyenne = round((note_devoir.note * 0.30) + (note_session.note * 0.70), 2)

            # Écrire les données avec alignement et bordures
            devoir_cell = sheet.cell(row=row_num, column=col_offset, value=note_devoir.note if note_devoir else 'N/A')
            devoir_cell.alignment = center_align
            devoir_cell.border = border
            
            session_cell = sheet.cell(row=row_num, column=col_offset + 1, value=note_session.note if note_session else 'N/A')
            session_cell.alignment = center_align
            session_cell.border = border
            
            moyenne_cell = sheet.cell(row=row_num, column=col_offset + 2, value=moyenne if moyenne is not None else 'N/A')
            moyenne_cell.alignment = center_align
            moyenne_cell.border = border
            
            col_offset += 3

    # Ajuster la largeur des colonnes
    column_widths = [15, 20, 20]  # Matricule, Nom, Prénom
    for matiere in matieres:
        column_widths.extend([12, 12, 12])  # Devoir, Session, Moyenne
        
    for col_num, width in enumerate(column_widths, 1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = width

    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="notes_{classe.nom.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    workbook.save(response)
    return response


def telecharger_notes_matiere_excel(request, classe_id, matiere_id):
    """
    Télécharge les notes d'une matière spécifique au format Excel avec un style épuré
    """
    classe = get_object_or_404(Classe, id=classe_id)
    matiere = get_object_or_404(Matiere_Programmer, id=matiere_id, classe=classe)
    transactions = Transaction.objects.filter(classe=classe).select_related('etudiant').order_by('etudiant__utilisateur__nom')

    # Créer le classeur Excel
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f'Notes - {matiere.nom}'

    # Définir la police pour l'entête
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # Définir les en-têtes
    headers = ["#", "Matricule", "Nom Prénom", f"{matiere.nom} (Devoir)", f"{matiere.nom} (Session)", f"{matiere.nom} (Moyenne)", "Statut"]
    
    # Écrire les en-têtes
    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Écrire les données
    for row_num, transaction in enumerate(transactions, 2):
        etudiant = transaction.etudiant
        
        # Récupérer les notes pour cette matière
        note_devoir = Notes.objects.filter(
            etudiant=etudiant, 
            matiere=matiere, 
            evaluation='devoir'
        ).first()
        note_session = Notes.objects.filter(
            etudiant=etudiant, 
            matiere=matiere, 
            evaluation='session'
        ).first()
        
        # Calculer la moyenne
        moyenne = None
        statut = "N/A"
        if note_devoir and note_session:
            moyenne = round((note_devoir.note * 0.30) + (note_session.note * 0.70), 2)
            statut = "Admis" if moyenne >= 10 else "Échec"

        # Remplir les infos de base
        sheet.cell(row=row_num, column=1, value=row_num-1).border = border  # Numéro de ligne
        sheet.cell(row=row_num, column=2, value=etudiant.utilisateur.matricule).border = border
        sheet.cell(row=row_num, column=3, value=f"{etudiant.utilisateur.nom} {etudiant.utilisateur.prenom}").border = border
        sheet.cell(row=row_num, column=4, value=note_devoir.note if note_devoir else 'N/A').border = border
        sheet.cell(row=row_num, column=5, value=note_session.note if note_session else 'N/A').border = border
        sheet.cell(row=row_num, column=6, value=moyenne if moyenne is not None else 'N/A').border = border
        sheet.cell(row=row_num, column=7, value=statut).border = border

    # Ajuster la largeur des colonnes
    column_widths = [5, 12, 25, 12, 12, 12, 12]
    for col_num, width in enumerate(column_widths, 1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = width

    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="notes_{matiere.nom.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    workbook.save(response)
    return response


def telecharger_notes_matiere_pdf(request, classe_id, matiere_id):
    """
    Télécharge les notes d'une matière spécifique au format PDF avec un style épuré
    """
    classe = get_object_or_404(Classe, id=classe_id)
    matiere = get_object_or_404(Matiere_Programmer, id=matiere_id, classe=classe)
    transactions = Transaction.objects.filter(classe=classe).select_related('etudiant').order_by('etudiant__utilisateur__nom')

    # Récupérer les notes pour cette matière
    etudiants_notes = []
    total_admis = 0
    total_echecs = 0
    somme_moyennes = 0
    compteur_moyennes = 0
    
    for transaction in transactions:
        etudiant = transaction.etudiant
        note_devoir = Notes.objects.filter(etudiant=etudiant, matiere=matiere, evaluation='devoir').first()
        note_session = Notes.objects.filter(etudiant=etudiant, matiere=matiere, evaluation='session').first()
        
        moyenne = None
        statut = "N/A"
        if note_devoir and note_session:
            moyenne = round((note_devoir.note * 0.30) + (note_session.note * 0.70), 2)
            if moyenne >= 10:
                statut = "Admis"
                total_admis += 1
            else:
                statut = "Échec"
                total_echecs += 1
            somme_moyennes += moyenne
            compteur_moyennes += 1
        else:
            statut = "-"

        etudiants_notes.append({
            'etudiant': etudiant,
            'note_devoir': note_devoir.note if note_devoir else None,
            'note_session': note_session.note if note_session else None,
            'moyenne': moyenne,
            'statut': statut
        })

    # Calculer les statistiques
    moyenne_generale = round(somme_moyennes / compteur_moyennes, 2) if compteur_moyennes > 0 else 0
    taux_reussite = round((total_admis / len(transactions)) * 100, 1) if transactions else 0

    context = {
        'classe': classe,
        'matiere': matiere,
        'etudiants_notes': etudiants_notes,
        'total_etudiants': len(transactions),
        'statistiques': {
            'admis': total_admis,
            'echecs': total_echecs,
            'moyenne_generale': moyenne_generale,
            'taux_reussite': taux_reussite
        },
        'date_generation': datetime.now().strftime('%d/%m/%Y à %H:%M'),
    }
    
    html_string = render_to_string('pages/pdf/notes_matiere_pdf.html', context)
    
    # Options pour un style propre
    options = {
        'page-size': 'A4',
        'margin-top': '0.75in',
        'margin-right': '0.75in',
        'margin-bottom': '0.75in',
        'margin-left': '0.75in',
        'encoding': "UTF-8",
        'no-outline': None,
        'enable-local-file-access': None,
        'disable-smart-shrinking': True,
    }
    
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result, encoding="utf-8", options=options)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="notes_{matiere.nom.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    return HttpResponse(f"Erreur de génération PDF: {pdf.err}")

def telecharger_notes_classe_pdf(request, classe_id):
    """
    Télécharge les notes de la classe au format PDF avec un style épuré
    """
    classe = get_object_or_404(Classe, id=classe_id)
    notes_par_matiere = menu_classe(request, classe_id, render_response=False)

    # Récupérer les statistiques générales
    total_etudiants = Transaction.objects.filter(classe=classe).count()
    matieres = Matiere_Programmer.objects.filter(classe=classe).distinct()
    
    context = {
        'classe': classe,
        'notes_par_matiere': notes_par_matiere,
        'matieres': matieres,
        'total_etudiants': total_etudiants,
        'date_generation': datetime.now().strftime('%d/%m/%Y à %H:%M'),
    }
    
    html_string = render_to_string('pages/pdf/notes_classe_pdf.html', context)
    
    # Options pour un style propre
    options = {
        'page-size': 'A4',
        'margin-top': '0.75in',
        'margin-right': '0.75in',
        'margin-bottom': '0.75in',
        'margin-left': '0.75in',
        'encoding': "UTF-8",
        'no-outline': None,
        'enable-local-file-access': None,
        'disable-smart-shrinking': True,
    }
    
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html_string.encode("UTF-8")), result, encoding="utf-8", options=options)

    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="notes_{classe.nom.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    return HttpResponse(f"Erreur de génération PDF: {pdf.err}")

@admin_decorators
def menu_classe(request, classe_id, render_response=True):
    # ... (le reste de la vue menu_classe reste inchangé)
    # Récupérer la classe
    classe = Classe.objects.get(id=classe_id)
    
    # Récupérer les étudiants de la classe via Transaction, triés par ordre alphabétique
    transactions = Transaction.objects.filter(
        classe=classe
    ).select_related('etudiant').order_by('etudiant__utilisateur__nom', 'etudiant__utilisateur__prenom')
    
    # Récupérer toutes les matières de la classe
    # matieres = EmploieDuTemps.objects.filter(
    # classe=classe,
    # ).select_related('matiere').values(
    # 'matiere__id',
    # 'matiere__nom', 
    # # 'matiere__code',
    # ).distinct()
    matieres = Matiere_Programmer.objects.filter(classe=classe).distinct()
    
    # Créer le dictionnaire des notes avec statistiques
    notes_par_matiere = {}
    
    for matiere in matieres:
        etudiants_notes = []
        total_admis = 0
        total_echecs = 0
        somme_moyennes = 0
        compteur_moyennes = 0
        
        for transaction in transactions:
            etudiant = transaction.etudiant
            
            # Récupérer les notes de devoir et session pour cette matière
            note_devoir = Notes.objects.filter(
                etudiant=etudiant,
                matiere=matiere,
                evaluation='devoir'
            ).first()
            
            note_session = Notes.objects.filter(
                etudiant=etudiant,
                matiere=matiere,
                evaluation='session'
            ).first()
            
            # Calculer la moyenne si les deux notes existent
            moyenne = None
            if note_devoir and note_session:
                moyenne = round((note_devoir.note * 0.30) + (note_session.note * 0.70), 2)
                somme_moyennes += moyenne
                compteur_moyennes += 1
                
                # Compter admis/échecs
                if moyenne >= 10:
                    total_admis += 1
                else:
                    total_echecs += 1
            
            etudiants_notes.append({
                'etudiant_id': etudiant.id,
                'nom': etudiant.utilisateur.nom,
                'prenom': etudiant.utilisateur.prenom,
                'matricule': etudiant.utilisateur.matricule, #Changer par le vrai matricule
                'note_devoir': note_devoir.note if note_devoir else None,
                'note_session': note_session.note if note_session else None,
                'moyenne': moyenne
            })
        
        # Calculer les statistiques finales
        moyenne_generale = round(somme_moyennes / compteur_moyennes, 2) if compteur_moyennes > 0 else 0
        taux_reussite = round((total_admis / len(etudiants_notes)) * 100, 1) if etudiants_notes else 0
        
        notes_par_matiere[matiere.id] = {
            'matiere_nom': matiere.nom,
            'matiere_code': getattr(matiere, 'code', 'N/A'),
            'etudiants_notes': etudiants_notes,
            'statistiques': {
                'total_etudiants': len(etudiants_notes),
                'admis': total_admis,
                'echecs': total_echecs,
                'moyenne_generale': moyenne_generale,
                'taux_reussite': taux_reussite
            }
        }
    if not render_response:
        return notes_par_matiere
    print(notes_par_matiere)
    context = {
        'classe': classe,
        'notes_par_matiere': notes_par_matiere,
        'matieres': matieres,
        'nombre_etudiants': transactions.count(),
    }
    
    return render(request,'pages/notes/notes_classe.html',context)


@admin_decorators
def menu(request):

    classes = Classe.objects.annotate(
        nombre_etudiants=Count('transaction__etudiant', distinct=True)
    ).order_by('nom')
    print (classes)
    
    context = {
        'etudiant':Transaction.objects.all(),
        'classes': classes
    }
    return render(request,'pages/notes/menu.html',context)

